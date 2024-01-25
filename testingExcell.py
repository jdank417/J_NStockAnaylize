import openpyxl
#import StockAnaylizerWithFMP as Stock


file_path = 'C:\\Users\\Jason Dank\\PycharmProjects\\FMPStockAnaylize\\standardDiv.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

#date_close_dict = Stock.get_close("AAPL", "2020-08-03", "2024-01-25")

# Print total number of rows and columns
#print('Total number of rows: ' + str(ws.max_row) + '. And total number of columns: ' + str(ws.max_column))

# Accessing and printing the value in cell B2
cell_value = ws['A2'].value

for row in ws.iter_rows(min_row=1, max_col=1, max_row=3, values_only=True):
    for cell in row:
        print(cell)

# Adding data to the next available row in column B
#for date, close_price in date_close_dict.items():
    #new_data = (f"Date: {date}, Close Price: {close_price}")
    #ws.append([new_data])

# Save the workbook
wb.save(file_path)
